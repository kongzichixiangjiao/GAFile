import openpyxl
from openpyxl.workbook.workbook import Workbook

path_source = "/Users/houjianan/Documents/GitHub/Python/Basic/Excel/source/"


class GAExcel(Workbook):

    def __init__(self, name):
        self.wb = self.loadExcel(name)

    def getFilepath(self, name):
        print("====")
        print(path_source + name)
        return path_source + name

    def loadExcel(self, name):
        print(name)
        path = self.getFilepath(name)
        return openpyxl.load_workbook(path)

    def sheetnames(self):
        return self.wb.sheetnames

    def getSH(self, name):
        return self.wb[name]

    def getRows(self, sh):
        return sh.rows

    def getCell(self, sh, row, column):
        return sh.cell(row, column)

    def getList(self, sh):
        return list(sh.rows)
