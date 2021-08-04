import openpyxl
import random

from openpyxl.workbook import workbook

""""
    读取excel文件
"""

import openpyxl

path = '/Users/houjianan/Documents/GitHub/Python/Basic/Excel/source/第二批车辆投放及商家入驻情况.xlsx'
wb = openpyxl.load_workbook(path)
# 获取所有工作表名
names = wb.sheetnames
# wb.get_sheet_by_name(name) 已经废弃,使用wb[name] 获取指定工作表
sheet = wb[names[0]]
# 获取最大行数
maxRow = sheet.max_row
# 获取最大列数
maxColumn = sheet.max_column
# 获取当前活动表
current_sheet = wb.active
# 获取当前活动表名称
current_name = sheet.title
# 通过名字访问Cell对象, 通过value属性获取值
a1 = sheet['A1'].value
# 通过行和列确定数据
a12 = sheet.cell(row=1, column=2).value
# 获取列字母
column_name = openpyxl.utils.cell.get_column_letter(1)
# 将列字母转为数字, 参数忽略大小写
column_name_num = openpyxl.utils.cell.column_index_from_string('a')
# 获取一列数据, sheet.iter_rows() 获取所有的行
"""
(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>)
(<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>)
(<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>)
(<Cell 'Sheet1'.A4>, <Cell 'Sheet1'.B4>, <Cell 'Sheet1'.C4>)
(<Cell 'Sheet1'.A5>, <Cell 'Sheet1'.B5>, <Cell 'Sheet1'.C5>)
"""
# print(sheet.iter_cols)
SN = []
index = 18010124000
min_row = 3
for one_column_data in sheet.iter_rows(min_row=min_row):
    A = one_column_data[0]
    if isinstance(A.value, int):
        A1 = str(A.value)[-3:]
        # print(A1)
        index += 1000
        item = index + int(A1)
        # print(item)
        sheet.cell(row=min_row, column=4, value=item)
        min_row += 1
    else:
        print("==")

wb.save(path)
print("保存完毕")

# for row_index, row_item in enumerate(data):

#     for col_index, col_item in enumerate(row_item):
#         # 写入
#         sheet.cell(row=row_index+1, column=col_index+1, value=col_item)

# # 写入excel文件 如果path路径的文件不存在那么就会自动创建
# workbook.save(path)
# print('写入成功')

# 获取一行数据, sheet.iter_cols() 获取所有的列
"""
(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>)
(<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>)
(<Cell 'Sheet1'.C1>, <Cell 'Sheet1'.C2>, <Cell 'Sheet1'.C3>)
"""
# for one_row_data in sheet.iter_cols():
#     print(one_row_data[1].value, end="\t")

# print("row = {}, column = {}".format(maxRow, maxColumn))


# def write_to_excel(path: str, sheetStr, info, data):

#     #     实例化一个workbook对象
#     workbook = openpyxl.Workbook()
#     # 激活一个sheet
#     sheet = workbook.active
#     # 为sheet设置一个title
#     sheet.title = sheetStr

#     # 添加表头（不需要表头可以不用加）
#     data.insert(0, list(info))
#     # 开始遍历数组
# for row_index, row_item in enumerate(data):

#     for col_index, col_item in enumerate(row_item):
#         # 写入
#         sheet.cell(row=row_index+1, column=col_index+1, value=col_item)

# # 写入excel文件 如果path路径的文件不存在那么就会自动创建
# workbook.save(path)
# print('写入成功')


# if __name__ == '__main__':

#     # 数据结构1 path 文件的路径
#     path = r'第二批车辆投放及商家入驻情况.xls'
#     # 数据结构1Excel 中sheet 的名字
#     sheetStr = '车辆-修改'

#     info = ['name', 'age', 'address']
#     # 数据结构1数据
#     writeData = [['John Brown', 18, 'New York No. 1 Lake Park']]

#     # 执行
#     write_to_excel(path, sheetStr, info, writeData)
