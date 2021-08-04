from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl
import os

# 当前文件的路径
pwd = os.getcwd()
# 当前文件的父路径
father_path = os.path.abspath(os.path.dirname(pwd)+os.path.sep+".")
# 当前文件的前两级目录
grader_father = os.path.abspath(os.path.dirname(pwd)+os.path.sep+"..")

print(pwd)
print(father_path)
print(grader_father)

workbook = openpyxl.load_workbook(pwd + '/Excel/source/test1.xlsx')

# newSheet = workbook.create_chartsheet("newSheetName")
newSheet = workbook.create_sheet("newSheetName")

sheetnames = workbook.sheetnames
for sheet in sheetnames:
    print(sheet)

test_sheet1 = workbook.get_sheet_by_name("test_sheet1")
Sheet = workbook["Sheet"]
print(test_sheet1)
print(Sheet)

worksheet = workbook.active
print(worksheet)

max_row = worksheet.max_row
max_column = worksheet.max_column

print(max_row, max_column)

selectcell = worksheet['A2']
selectcell1 = worksheet.cell(row=1, column=3)
print(selectcell.row)
print(selectcell.column)
print(selectcell.coordinate)
print(selectcell1.value)

selectcol = worksheet["C"]
print(selectcol)
selectcell2 = selectcol[4]
print(selectcell2)
selectcols = worksheet["B:C"]
print(selectcols)
selectcolsC = selectcols[1]
print(selectcolsC)

allcol = worksheet.columns
print(allcol)

selectrow = worksheet[2]
print(selectrow)
selectrowA = selectrow[1]
print(selectrowA)


selectrows = worksheet["1:4"]
print(selectrows)

allrow = worksheet.rows
print(allrow)

cell_range = worksheet["A1:C3"]
print(cell_range)

for row_range in cell_range:
    for cell in row_range:
        print(cell.value)

print(get_column_letter(2))  # 2 => B
print(column_index_from_string("C"))  # C => 3
