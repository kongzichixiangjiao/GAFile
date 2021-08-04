import os
from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl
import openpyxl

workbook = openpyxl.Workbook()

workbook.create_sheet(index=1, title="第二张表")

sheet = workbook.active
sheet.title = "设置的表单名称"
sheet["A1"] = "New Value"

workbook.create_sheet(index=2, title="Sheet3")
sheet3 = workbook["Sheet3"]
workbook.remove(sheet3)


# 批量写入数据
# 方法一
# workbook = openpyxl.Workbook()
ws1 = workbook.create_sheet("第一页")
for row in range(40):
    ws1.append(range(17))


# 方法二
# workbook = openpyxl.Workbook()
ws2 = workbook.create_sheet("第二页")
rows = [
    ["Number", "Batch1", "Batch2"],
    [2, 40, 30], [3, 50, 25],
    [4, 30, 30], [5, 60, 10]
]
for row in rows:
    ws2.append(row)


# 方法三
# workbook = openpyxl.Workbook()
ws3 = workbook.create_sheet("第三页")
for row in range(5, 30):
    for col in range(15, 24):
        ws3.cell(column=col, row=row, value=123)

ws4 = workbook.create_sheet("第四页")
ws4.insert_cols(idx=2)


# 当前文件的路径
pwd = os.getcwd()

# 修改完毕保存到文件
workbook.save(pwd + '/Excel/source/created.xlsx')
