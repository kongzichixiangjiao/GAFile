import openpyxl

path = './第二批车辆投放及商家入驻情况.xlsx'
wb = openpyxl.load_workbook(path)
# 获取所有工作表名
names = wb.sheetnames
# wb.get_sheet_by_name(name) 已经废弃,使用wb[name] 获取指定工作表
sheet = wb[names[0]]

SN = []
index = 18010259000
min_row = 3
for one_column_data in sheet.iter_rows(min_row=min_row):
    A = one_column_data[0]
    if isinstance(A.value, int) | isinstance(A.value, str):
        A1 = str(A.value)[-3:]
        # print(A1)
        index += 1000
        item = index + int(A1)
        # print(item)
        sheet.cell(row=min_row, column=4, value=item)
        min_row += 1
    else:
        print("==出错了...")

wb.save(path)

print("生成完毕-已保存")
