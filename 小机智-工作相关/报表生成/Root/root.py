import openpyxl
import time

__all__ = ['dates', 'stores']

localtime = time.localtime()

year = localtime.tm_year
month = localtime.tm_mon
day = localtime.tm_mday
print(year, month, day)

path = '../车电一体台账清单-'+str(year)+'年'+str(month).zfill(2)+'月'+str(day)+'日.xlsx'
print(path)

wb = openpyxl.load_workbook(path)
# 获取所有工作表名
names = wb.sheetnames
print(names)
sheet = wb[names[0]]
print(sheet.title)

print(sheet.iter_rows)
print(sheet.iter_rows(min_row=2))

stores = []
dates = []
min_row = 2
for one_column_data in sheet.iter_rows(min_row=min_row):
    agent = one_column_data[1]
    store = one_column_data[2]
    date = one_column_data[11]
    print(store.value)
    print(date.value)
    print(type(date.value))
    rDate = str(year)+str(month).zfill(2)+str(day)
    # if date.value.find(rDate) != -1:
    print("===")
    print(agent.value)
    if agent.value == '华夏成都测试门店':
        stores.append(store.value)
        dates.append(date.value)


print("数据源：")
print(len(dates))

wb.close()
