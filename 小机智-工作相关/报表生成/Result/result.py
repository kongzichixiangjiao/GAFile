
import sys
import os
import openpyxl
import time

import_dir = \
    os.path.join(os.path.join(os.path.dirname(
        __file__), os.pardir), 'Root')
sys.path.insert(0, import_dir)
file_name = 'root'
Root = __import__(file_name)
print(Root.dates)
print(Root.stores)

import_dir1 = \
    os.path.join(os.path.join(os.path.dirname(
        __file__), os.pardir), 'Tool')
sys.path.insert(0, import_dir1)
file_name1 = 'tool'
Tool = __import__(file_name1)


path = '../成都铁塔平台车辆日报表.xlsx'

Tool.sameExcelCopySheet(path, '6.14', '6.15')

wb = openpyxl.load_workbook(path)

sheet = wb['6.15']

# localtime = time.localtime()
# month = localtime.tm_mon
# day = localtime.tm_mday

# sheet = wb.create_sheet(str(month) + '.' + str(day).zfill(2))


targetStores = []
min_row = 3
for one_column_data in sheet.iter_rows(min_row=min_row):
    # 店名称
    store = one_column_data[1].value
    targetStores.append(store)
    num = Root.stores.count(store)
    # 第一批投放量
    a1 = one_column_data[3].value
    # 第二批投放量
    a2 = one_column_data[4].value
    # 已出租数量
    czNum = one_column_data[8].value
    # 滞留数量
    zlNum = one_column_data[10].value
    # 第一批投放日期
    a1_date = one_column_data[6].value
    # 第二批投放日期
    a2_date = one_column_data[7].value
    lastDate = ''

    if a1_date is None:
        lastDate = str(a2_date)
    else:
        if a2_date is None:
            lastDate = str(a1_date)
        else:
            lastDate = str(a2_date)

    localtime = time.localtime()

    year = localtime.tm_year
    month = localtime.tm_mon
    day = localtime.tm_mday

    # 滞留天数
    # if lastDate == 'None':
    #     print("lastDate is None")
    # else:
    #     if czNum != 0:
    #         d = Tool.Caltime('2021.' + lastDate, '2021.' +
    #                          str(month).zfill(2) + '.' + str(day))
    #         print("已出租天数")
    #         print(d.days)
    #         print(type(print(d.days)))
    #         sheet.cell(column=12, row=min_row, value=d.days)
    #     else:
    #         print(0)

    if isinstance(a1, int):
        if czNum != num:
            sheet.cell(column=9, row=min_row, value=num)
            sheet.cell(column=11, row=min_row, value=a1 + a2 - num)

    if store in Root.stores:
        print("")
    else:
        if isinstance(store, str):
            print("ERROR......." + store)

    min_row += 1

isSave = True

for item in Root.stores:
    if item in targetStores:
        print('')
    else:
        print("ERROR-------有新店铺需要手动新增")
        isSave = False
        break
if isSave:
    print("保存成功")
    wb.save(path)
