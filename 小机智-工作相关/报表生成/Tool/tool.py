# python3读写excel
'''
实现功能：
要把test1中的sheet表内容复制到test表sheet1中
'''

from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
import os
import time
import datetime


def sameExcelCopySheet(ExcelPath, sourceSheetName, targetSheetName):
    # 读取数据
    # wb1 = openpyxl.load_workbook('test1.xlsx')
    wb1 = openpyxl.load_workbook(ExcelPath)
    if targetSheetName in wb1.sheetnames:
        print("WARNNING--------已经存在：" + targetSheetName)
    else:
        sheets1 = wb1.get_sheet_names()  # 获取sheet页
        sheets2 = wb1.get_sheet_names()
        sheet1 = wb1.get_sheet_by_name(sourceSheetName)
        sheet2 = wb1.create_sheet(targetSheetName)

        for i, row in enumerate(sheet1.iter_rows()):
            for j, cell in enumerate(row):
                sheet2.cell(row=i+1, column=j+1, value=cell.value)

        wb1.save(ExcelPath)  # 保存数据
        wb1.close()  # 关闭excel
        print("-------------成功复制------------")


# sameExcelCopySheet('../成都铁塔平台车辆日报表.xlsx', '6.14', '6.15')


# 判断日期是否为合法输入，年月日的格式需要与上面对应，正确返回True，错误返回False，注意大小写。
def is_date(str):
    try:
        time.strptime(str, "%Y-%m-%d")
        return True
    except:
        return False


# 计算两个日期相差天数，自定义函数名，和两个日期的变量名。
def Caltime(date1, date2):
    # %Y-%m-%d为日期格式，其中的-可以用其他代替或者不写，但是要统一，同理后面的时分秒也一样；可以只计算日期，不计算时间。
    #date1=time.strptime(date1,"%Y-%m-%d %H:%M:%S")
    #date2=time.strptime(date2,"%Y-%m-%d %H:%M:%S")
    date1 = time.strptime(date1, "%Y.%m.%d")
    date2 = time.strptime(date2, "%Y.%m.%d")
    # 根据上面需要计算日期还是日期时间，来确定需要几个数组段。下标0表示年，小标1表示月，依次类推...
    # date1=datetime.datetime(date1[0],date1[1],date1[2],date1[3],date1[4],date1[5])
    # date2=datetime.datetime(date2[0],date2[1],date2[2],date2[3],date2[4],date2[5])
    date1 = datetime.datetime(date1[0], date1[1], date1[2])
    date2 = datetime.datetime(date2[0], date2[1], date2[2])
    # 返回两个变量相差的值，就是相差天数
    return date2-date1


# print(Caltime('2021.12.12', '2021.12.15'))
# print(type(Caltime('2021.12.12', '2021.12.15')))
# print(Caltime('2021.12.12', '2021.12.15').days)

# print(Caltime('2021.02.12', '2021.2.15'))
# print(type(Caltime('2021.12.12', '2021.12.15')))
# print(Caltime('2021.12.12', '2021.12.15').days)
